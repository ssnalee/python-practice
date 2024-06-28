from openpyxl import Workbook
from openpyxl import load_workbook #파일 불러오기
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.drawing.image import Image
from random import *
from openpyxl.chart import BarChart,Reference,LineChart
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
# wb = Workbook() #새 워크북 생성

##1 

# ws = wb.active #현재 활성화된 sheet 가져옴
# ws.title = "NadoSheet"

##2
# ws = wb.create_sheet() #새로운 시트 기본이름으로 생성
# ws.title = 'MySheet' 
# ws.sheet_properties.tabColor = 'ff66ff'

# ws1 = wb.create_sheet('yourSheet') #주어진 이름으로 시트생성
# ws2 = wb.create_sheet('newSheet',2) #두번째 인덱스에 시트 생성
# new_ws = wb["newSheet"] #Dict 형태로 sheet에 접근
# print(wb.sheetnames) #모든 시트이름 확인

# #sheet 복사
# new_ws['A1'] = "Test"
# target = wb.copy_worksheet(new_ws)
# target.title = 'Copied Sheet'

##3
# ws = wb.active
# ws.title = 'NadoSheet'

# #A1 셀에 1이라는 값 입력
# ws["A1"] = 1
# ws["A2"] = 2
# ws["A3"] = 3
# ws["B1"] = 4
# ws["B2"] = 5
# ws["B3"] = 6

# print(ws['A1']) #A1 셀의 정보를 출력
# print(ws['A1'].value) #A1 셀의 '값'을 출력
# print(ws['A10'].value) #값이 없을 땐 'None' 출력

# print(ws.cell(column=1,row=1).value) #ws['A1'].value
# print(ws.cell(column=2,row=1).value) #ws['B1'].value

# c = ws.cell(column=3, row=1, value=10) #ws['C1'].value = 10
# print(c.value)  #ws['C1'].value
# #반복문을 이용해서 랜덤 숫자 채우기
# index = 1
# for x in range(1,11) :
#     for y in range(1, 11) :
#         ws.cell(row = x, column= y,value=index)
#         index +=1

# wb.save('sample.xlsx')


##4
# wb = load_workbook('sample.xlsx') #sample.xlsx 파일에서 wb를 불러옴
# ws = wb.active #활성화된 sheet

# #cell 데이터 불러오기
# for x in range(1,11) :
#     for y in range(1,11) :
#         print(ws.cell(row=x,column=y).value, end = " ")
#     print()
    
# #cell 갯수를 모를 때
# for x in range(1, ws.max_row +1) : 
#     for y in range(1,ws.max_column +1) :
#         print(ws.cell(row=x,column=y).value, end =" ")
#     print()

##5
# ws = wb.active

# ws.append(['번호','영어','수학']) #A,B,C
# for i in range(1,11) : #10개 데이터 넣기
#     ws.append([i,randint(0,100),randint(0,100)])

# col_B = ws['B'] #영어 column 만 가지고오기
# print(col_B)
# for cell in col_B :
#     print(cell.value)

# col_range = ws ['B:C']
# for cols in col_range :
#     for cell in cols :
#         print(cell.value)

# row_title = ws[1] #1번째 row만 가지고 오기
# for cell in row_title :
#     print(cell.value, end=" ")
# print()
# row_range = ws[2:6] #2번쨰 줄에서 6번째 줄까지 가지고오기
# for rows in row_range :
#     for cell in rows :
#         print(cell.value, end=" ")
#     print()

# row_range = ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지
# for rows in row_range :
#     for cell in rows :
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ") # A/10
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end = " ")
#         print(xy[0],end=" ") #A
#         # print(xy[1],end=" ") #1
#     print()
    
#전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows) :
#     print(row[2].value)

#전체 columns
# print(tuple(ws.columns))
# for column in tuple(ws.columns) :
#      print(column[0].value)

# for row in ws.iter_rows() : #전체 row
#     print(row[1].value)
# for column in ws.iter_cols() : #전체 columns
#     print(column[1].value)

# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지
# for row in ws.iter_rows( min_row=1, max_row=11 , min_col=2, max_col=3) : 
#     print(row[1].value) #수학, 영어
#     print(row)

# for col in ws.iter_cols(min_row=1,max_row=5,min_col=1,max_col=3) :
#     print(col[2].value)

# wb.save('sample.xlsx')


##6
# wb = load_workbook('sample.xlsx')
# ws = wb.active

# for row in ws.iter_rows(min_row=2) :
#     # 번호, 영어, 수학
#     if int(row[1].value) > 80 :
#         print(row[0].value,'번 학생은 영어 천재')
        
# for row in ws.iter_rows(max_row =1) :
#     for cell in row :
#         if cell.value == '영어' :
#             cell.value = '컴퓨터'
            
# wb.save('sample_modified.xlsx')

# ws.insert_rows(8,5) #8번째 위치에 5줄 추가
# ws.insert_rows(8) #8번줄이 비워짐
# ws.insert_cols(2) #B번째 열이 비워짐 (새로운 열 추가)
# ws.insert_cols(2,3) #B번째 열로부터 3열 추가
# wb.save('sample_insert.xlsx')
# ws.delete_rows(8) #8번째 줄에 있는 7번학생 데이터 삭제
# ws.delete_rows(8,3) #8번째 줄부터 총 3줄 삭제
# ws.delete_cols(1) #1번째 열(A)삭제
# ws.delete_cols(1,2) #1번째 열(A)부터 총 2줄 삭제
# wb.save('sample_del.xlsx')

#번호 영어 수학
#번호 (국어) 영어 수학

# ws.move_range('B1:C11',rows=0,cols=1)
# ws['B1'].value = '국어' #B1셀에 국어 입력

# ws.move_range('C1:C11',rows=0,cols=-1)
# wb.save('sample_ko.xlsx')


##7

# B2:C11 까지의 데이터를 차트로 생성
# bar_value = Reference(ws, min_row=2,max_row=11,min_col=2,max_col=3)
# bar_chart = BarChart() #차트 종류 설정
# bar_chart.add_data(bar_value) #차트 데이터 추가
# ws.add_chart(bar_chart,'E1')

# line_value = Reference(ws, min_row=2,max_row=11,min_col=2,max_col=3)
# linechart = LineChart()
# linechart.add_data(line_value, titles_from_data=True) #계열 > 영어,수학 (제목에서 가져옴)
# linechart.title = '성적표'
# linechart.style = '20'
# linechart.y_axis.title = '점수' #Y축 제목
# linechart.x_axis.title = '번호' #X축 제목
# ws.add_chart(linechart,'E1')

# wb.save('sample_chart.xlsx')

##8
#번호, 영어, 수학
# a1 = ws['A1'] #번호
# b1 = ws['B1'] #영어
# c1 = ws['C1'] #수학

# # A 열의 너비를 5로 설정
# ws.column_dimensions['A'].width = 5

# # 1행의 높이를 50으로 설정
# ws.row_dimensions[1].height = 50

# # 스타일 적용
# a1.font = Font(color='FF0000',italic=True, bold=True) #글자색 빨갛게, 이탤릭, 두껍게 적용
# b1.font = Font(color='CC33FF',name="Arial",strike=True) #폰트를 Arial, 취소선 추가
# c1.font = Font(color='0000FF',size=20,underline='double') #폰트를 Arial, 취소선 추가

# # 테두리 적용
# thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
# a1.border = thin_border
# b1.border = thin_border
# c1.border = thin_border

# # 90점 넘는 셀에 대해서 초록색으로 적용
# for row in ws.rows :
#     for cell in row :
#         cell.alignment = Alignment(horizontal='center',vertical='center') #정렬 center,left,right,top,bottom
#         if cell.column == 1 :
#             continue
#         if isinstance(cell.value,int) and cell.value > 90 :
#             cell.fill = PatternFill(fgColor='00FF00', fill_type='solid')
#             cell.font = Font(color="FF0000")
# # 틀 고정
# ws.freeze_panes = 'B2' # B2기준으로 틀 고정

# wb.save('sample_style.xlsx')

import datetime

##9
wb = Workbook()
ws = wb.active

# ws['A1'] = datetime.datetime.today() #오늘 날짜 정보
# ws['A2'] = '=SUM(1,2,3)'
# ws['A3'] = '=AVERAGE(1,2,3)'
# ws['A4'] = 10
# ws['A5'] = 20
# ws['A6'] = '=SUM(A4:A5)'

# wb.save('sample_formula.xlsx')
# wb = load_workbook('sample_formula.xlsx',data_only=True)
# ws = wb.active
# # 수식 그대로 가져오고 있음
# # evaluate 되지 않은 상태의 데이터는 None 이라고 표시 파일들어가서 한번 저장해야 데이터 다 들어옴
# for row in ws.values :
#     for cell in row :
#         print(cell)


##10

#병합하기
# ws.merge_cells('B2:D2')
# ws['B2'].value = 'merge cell'
# wb.save('sample_merge.xlsx')

# wb = load_workbook('sample_merge.xlsx',data_only=True)
# ws = wb.active

# #병합해제
# ws.unmerge_cells('B2:D2')
# wb.save('sample_unmerge.xlsx')

# img = Image('img.png')

# # C3위치에 img.png 파일의 이미지를 삽입
# ws.add_image(img,"C3")
# wb.save('sample_image.xlsx')

# ImportError : You must install Pillow to fetch image....


##QUIZ

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.append(("학번","출석","퀴즈1","퀴즈2","중간고사","기말고사","프로젝트"))
scores = [
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20),
]

for s in scores :
    ws.append(s)
    
for idx, cell in enumerate("D") :
    if idx == 0 :
        continue
    cell.value = 10
    
ws['H1'] = '총점'
ws['I1'] = '성적'

for idx, score in enumerate(scores,start =2) :
    sum_val = sum(score[1:]) - score[3] + 10 #총점
    ws.cell(row=idx,column=8).value = "=SUM(B{}:G{})".format(idx,idx)
    grade = None
    if sum_val >= 90 :
        grade ="A"
    elif sum_val >= 80 :
        grade ="B"    
    elif sum_val >= 70 :
        grade ="C" 
    else :
        grade ="D" 
    if score[1] <5 :
        grade = "F"
    ws.cell(row=idx,column=9).value = grade
wb.save('scores.xlsx')
